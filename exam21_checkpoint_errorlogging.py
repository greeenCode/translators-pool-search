# -*- coding: utf-8 -*-

## 소스 폴더 내 pdf, html에서 텍스트를 추출하여 OpenAI api 호출, 분석하여 번역사 정보를 엑셀파일에 저장 ##
## 다수 파일처리를 위한 
## Checkpoint 저장 및 재개 기능 추가 - 
## 에러 처리 및 로깅 기능 추가 - 에러 발생한 파일 정보를 기록 후 계속 진행

import os
import pandas as pd
from datetime import datetime
import openai
import yaml
import fitz  # PyMuPDF
import json
from openpyxl import Workbook, load_workbook
from html_to_txt_return import extract_text_from_html
import re
import pickle

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Configurations
source_folder = r'abba'
target_folder = r'abba\@extracted'
target_path = os.path.join(target_folder, 'file_info12.xlsx')
checkpoint_path = os.path.join(target_folder, 'checkpoint.pkl')
log_path = os.path.join(target_folder, 'error_log.txt')

# Load API Key from credentials.yml
with open('config/credentials.yml', 'r') as file:
    credentials = yaml.safe_load(file)
api_key = credentials['openai']['api_key']
openai.api_key = api_key

# Helper functions
def extract_text_from_pdf(file_path):
    text = ""
    try:
        with fitz.open(file_path) as doc:
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                text += page.get_text("text")
    except Exception as e:
        log_error(f"Error extracting text from PDF {file_path}: {e}")
    return text

def extract_text_from_txt(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            text = file.read()
    except Exception as e:
        log_error(f"Error extracting text from TXT {file_path}: {e}")
        text = ""
    return text

def extract_text_from_html(file_path):
    text = ""
    try:
        # Selenium 옵션 설정
        chrome_options = Options()
        chrome_options.add_argument("--headless")  # 브라우저 창을 띄우지 않음
        # chrome_options.add_argument("--disable-gpu")  # GPU 사용 안함
        chrome_options.add_argument("--no-sandbox")  # 샌드박스 모드 사용 안함

        # ChromeDriver 경로 설정
        chrome_driver_path = r'C:\Util\chromedriver-win64\chromedriver.exe'  # ChromeDriver 경로로 변경
        service = Service(chrome_driver_path)

        # 브라우저 열기
        driver = webdriver.Chrome(service=service, options=chrome_options)

        try:
            # HTML 파일 열기
            driver.get(f'file:///{os.path.abspath(file_path)}')
            
            # 잠시 대기하여 페이지가 완전히 로드되도록 함
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
            
            # 페이지의 전체 텍스트 추출
            text = driver.find_element(By.TAG_NAME, 'body').text
            
        except Exception as e:
            log_error(f"Error extracting text from HTML {file_path}: {e}")
                       
        finally:
            # 드라이버 종료
            driver.close()
            driver.quit()
            
    except Exception as e:
        log_error(f"Error initializing WebDriver: {e}")
    
    return text

def format_multiline_text(text):
    if isinstance(text, list):
        text = '; '.join(text)
    return text.replace('; ', ';\n')

def clean_response_text(text):
    text = re.sub(r'[\x00-\x1F\x7F]', '', text)
    return text

def extract_information(text):
    prompt_text = f"""
    주어진 텍스트를 분석해서 다음 정보 항목을 JSON 형식으로 추출해줘. 
    번역사의 이름, 이메일, 전화번호, 현 거주지(도시 이름까지만), 자기 소개 개요, 번역 경력년수(명시되어있다면 명시된 년수와 활동 내역으로 추정한 시작년도를 표기하고, 명시되지 않았다면 활동 내역으로 추정한 시작년도부터 현재까지의 경과년수와 시작년도를 표기해줘. 알 수 없다면 '알 수 없음'), 번역 가능한 언어, 통역 가능한 언어, 번역 툴  사용가능여부(Trados, MemoQ, Smartcat 등 번역 툴 사용가능하면 툴 이름을 표기하고, 알수없으면 "알 수 없음"), 주요 학력, 주요 경력, 해외(한국 외) 교육기관에서 공부경험 유무, 그밖에 번역사로서 경쟁력 등을 아래의 출력문 사례처럼 작성해줘.

    {{
        "이름": "양중남",
        "이메일": "inpraiseofdissent@gmail.com",
        "전화번호": "010-8751-1205",
        "거주지": "제주시",
        "자기소개": "양중남은 한영 통역 대학원을 졸업하고 미국에서 심리학 박사 학위를 받은 후 20년간 연구원으로 재직하였으며, 현재는 8년 경력의 프리랜서 번역사로 활동하고 있습니다. 영어와 한국어를 자유롭게 구사하는 바이링구얼 번역사입니다.",
        "경력년수": "8년 from 2001",
        "번역가능언어": "영어>한국어, 한국어>영어",
        "통역가능언어": "영어, 일본어",
        "번역툴가능여부": Trados, MemoQ
        "주요학력": 
        "박사, 실험 심리학, New York University (1995-1999); 
        학사, 영어 교육과, 제주 대학교 (1977-1981)",
        "주요경력": 
        "프리랜서 번역사 (2012-현재); 제주 대학 교육학과 대학원 강사 (2016-2018); 
        연구원, NASA Ames 연구소 (2002-2004); 박사후 과정, University of Chicago (1999-2001)",
        "해외학업유무": "New York University, Ball State University, University of Chicago",
        "경쟁력": 
        "다양한 분야 번역 경험 (자연과학, 사회과학, 비즈니스, 금융, 의학, 컴퓨터 과학 및 IT); 
        영한 자막 번역 경험 (의학, 교육, 음악, 드라마 등 다양한 분야에서 약 700개 비디오 번역); 
        미국에서 다수의 연구 논문 발표"
    }}

    텍스트:
    {text}
    """
    response = openai.ChatCompletion.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": "You are a data extraction assistant."},
            {"role": "user", "content": prompt_text}
        ],
        max_tokens=1500,
        temperature=0.5
    )
    response_text = response['choices'][0]['message']['content'].strip()
    
    # Clean the response text
    response_text = response_text.replace("```json", "").replace("```", "").strip()
    response_text = clean_response_text(response_text)

    try:
        extracted_info = json.loads(response_text)
        if '주요학력' in extracted_info:
            extracted_info['주요학력'] = format_multiline_text(extracted_info['주요학력'])
        if '주요경력' in extracted_info:
            extracted_info['주요경력'] = format_multiline_text(extracted_info['주요경력'])
        if '경쟁력' in extracted_info:
            extracted_info['경쟁력'] = format_multiline_text(extracted_info['경쟁력'])
        if '해외학업유무' in extracted_info:
            extracted_info['해외학업유무'] = format_multiline_text(extracted_info['해외학업유무'])
        return extracted_info
    except json.JSONDecodeError as e:
        log_error(f"Error parsing JSON: {e}\nResponse text: {response_text}")
        return None

def log_error(message):
    with open(log_path, 'a') as log_file:
        log_file.write(f"{datetime.now().isoformat()} - {message}\n")

# Checkpoint loading
if os.path.exists(checkpoint_path):
    with open(checkpoint_path, 'rb') as f:
        checkpoint_data = pickle.load(f)
        processed_files = checkpoint_data.get('processed_files', [])
else:
    processed_files = []

# Main script
file_data = []
system_files = ['desktop.ini', 'Thumbs.db']
excluded_char = '@'
default_file_date = datetime(2000, 1, 1)

use_default_date = input(f"Do you want to use the default file date {default_file_date.date()}? (yes/y to use default): ").strip().lower()
if use_default_date in ('yes', 'y', ''):
    modified_file_date = default_file_date
else:
    date_input = input("Enter the modified file date (YYYY-MM-DD): ").strip()
    try:
        modified_file_date = datetime.strptime(date_input, '%Y-%m-%d')
    except ValueError:
        print("Invalid date format. Please enter the date in YYYY-MM-DD format.")
        exit()

file_count = 0
for root, dirs, files in os.walk(source_folder):
    dirs[:] = [d for d in dirs if excluded_char not in d]
    files = [file for file in files if file not in system_files and excluded_char not in file and (file.lower().endswith('.pdf') or file.lower().endswith('.html'))]
    for file in files:
        file_path = os.path.join(root, file)
        file_modified_time = datetime.fromtimestamp(os.path.getmtime(file_path))
        if file_modified_time > modified_file_date:
            file_count += 1

print(f"Total number of files in '{source_folder}' modified after {modified_file_date.date()}: {file_count}")
proceed = input("Do you want to proceed? (yes/y to continue): ").strip().lower()

if proceed in ('yes', 'y', ''):
    processed_count = 0
    for root, dirs, files in os.walk(source_folder):
        dirs[:] = [d for d in dirs if excluded_char not in d]
        files = [file for file in files if file not in system_files and excluded_char not in file and (file.lower().endswith('.pdf') or file.lower().endswith('.html'))]
        for file in files:
            file_path = os.path.join(root, file)
            if file_path in processed_files:
                continue  # Skip already processed files

            file_modified_time = datetime.fromtimestamp(os.path.getmtime(file_path))
            if file_modified_time > modified_file_date:
                processed_count += 1
                print(f"{processed_count}/{file_count} 번째 파일 작업 중 ... (파일명: {file})")

                try:
                    if file.lower().endswith('.pdf'):
                        text = extract_text_from_pdf(file_path)
                    elif file.lower().endswith('.html'):
                        text = extract_text_from_html(file_path)
                    else:
                        continue
                    extracted_info = extract_information(text)
                    if extracted_info is not None:
                        extracted_info['파일수정일'] = file_modified_time.strftime('%Y-%m-%d')
                        relative_file_path = os.path.relpath(file_path, start=target_folder)
                        extracted_info['File Link'] = f'=HYPERLINK("{relative_file_path}")'
                        file_data.append(extracted_info)

                        # Add to processed files and save checkpoint
                        processed_files.append(file_path)
                        with open(checkpoint_path, 'wb') as f:
                            pickle.dump({'processed_files': processed_files}, f)
                except Exception as e:
                    log_error(f"Error processing file {file_path}: {e}")

    if file_data:
        if os.path.exists(target_path):
            wb = load_workbook(target_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            headers = ["이름", "File Link", "이메일", "전화번호", "거주지", "자기소개", "경력년수", "번역가능언어", "통역가능언어", "번역툴가능여부", "주요학력", "주요경력", "해외학업유무", "경쟁력", "파일수정일"]
            ws.append(headers)
        
        next_row = ws.max_row + 1
        
        for info in file_data:
            row = [
                info.get("이름", ""),
                info.get("File Link", ""),
                info.get("이메일", ""),
                info.get("전화번호", ""),
                info.get("거주지", ""),
                info.get("자기소개", ""),
                info.get("경력년수", ""),
                info.get("번역가능언어", ""),
                info.get("통역가능언어", ""),
                info.get("번역툴가능여부", ""),
                format_multiline_text(info.get("주요학력", "")),
                format_multiline_text(info.get("주요경력", "")),
                format_multiline_text(info.get("해외학업유무", "")),
                format_multiline_text(info.get("경쟁력", "")),
                info.get("파일수정일", "")
            ]
            ws.append(row)
       
        os.makedirs(target_folder, exist_ok=True)
        try:
            wb.save(target_path)
            print(f"Finished! File information saved to '{target_path}'")
        except Exception as e:
            log_error(f"Error saving Excel file: {e}")
    else:
        print("No valid files processed.")
else:
    print("Process aborted by the user.")
